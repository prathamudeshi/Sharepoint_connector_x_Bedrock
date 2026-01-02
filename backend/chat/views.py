from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.conf import settings
from connectors.gemini_service import GeminiService
from connectors.sharepoint_service import SharePointService
import logging

logger = logging.getLogger(__name__)

from rest_framework import status, permissions, authentication
from accounts.models import SharePointCredentials

# ... imports ...

class ChatView(APIView):
    authentication_classes = [authentication.TokenAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        """
        Payload:
        {
            "message": "User query",
            "history": [...], 
            "context_files": ["file_content_or_ref"]
        }
        """
        try:
            message = request.data.get("message")
            history = request.data.get("history", [])
            context_files = request.data.get("context_files", [])

            if not message:
                return Response({"error": "Message is required"}, status=status.HTTP_400_BAD_REQUEST)

            # Initialize Service
            llm_service = GeminiService()
            
            # Fetch content for selected files
            full_context = []
            if context_files:
                try:
                    sp_service = SharePointService(user=request.user)
                    
                    import io
                    import mimetypes

                    # Ensure common types are recognized
                    mimetypes.add_type("application/vnd.openxmlformats-officedocument.wordprocessingml.document", ".docx")
                    mimetypes.add_type("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", ".xlsx")
                    mimetypes.add_type("application/pdf", ".pdf")

                    for file_obj in context_files:
                        # Handle both string (legacy/fallback) and object formats
                        if isinstance(file_obj, str):
                             logger.warning(f"Received string filename: {file_obj}. Expected object with ID/Url.")
                             continue
                        
                        file_name = file_obj.get('name')
                        file_id = file_obj.get('id')
                        download_url = file_obj.get('downloadUrl')
                        
                        if not file_id and not download_url:
                             continue

                        try:
                            content = sp_service.get_file_content(file_id=file_id, download_url=download_url)
                        except Exception as e:
                            logger.error(f"Download error for {file_name}: {e}")
                            full_context.append(f"Error downloading {file_name}: {e}")
                            content = None

                        if content:
                            lower_name = file_name.lower() if file_name else ""
                            mime_type, _ = mimetypes.guess_type(file_name)
                            
                            # 1. Handle Office Documents (Text Extraction)
                            # Gemini does not natively support DOCX/XLSX as binary parts, so we must extract text.
                            if lower_name.endswith('.docx'):
                                try:
                                    from docx import Document
                                    doc = Document(io.BytesIO(content))
                                    text_content = "\n".join([para.text for para in doc.paragraphs])
                                    full_context.append(f"Filename: {file_name} (Extracted Content)\n{text_content}")
                                except Exception as docx_err:
                                    full_context.append(f"Error reading DOCX {file_name}: {docx_err}")
                                    
                            elif lower_name.endswith('.xlsx'):
                                try:
                                    import openpyxl
                                    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
                                    text_content = ""
                                    for sheet in wb.sheetnames:
                                        text_content += f"Sheet: {sheet}\n"
                                        ws = wb[sheet]
                                        for row in ws.iter_rows(values_only=True, max_row=50):
                                            text_content += "\t".join([str(c) if c is not None else "" for c in row]) + "\n"
                                    full_context.append(f"Filename: {file_name} (Extracted Content)\n{text_content}")
                                except Exception as xlsx_err:
                                    full_context.append(f"Error reading XLSX {file_name}: {xlsx_err}")

                            # 2. Handle Native Supported Types (PDF, Images)
                            # Gemini supports valid image formats and PDF natively.
                            elif lower_name.endswith('.pdf') or (mime_type and mime_type.startswith('image/')):
                                if not mime_type:
                                    # Fallback for PDF if guess failed
                                    if lower_name.endswith('.pdf'):
                                        mime_type = 'application/pdf'
                                    elif lower_name.endswith('.jpg') or lower_name.endswith('.jpeg'):
                                        mime_type = 'image/jpeg'
                                    elif lower_name.endswith('.png'):
                                        mime_type = 'image/png'
                                    elif lower_name.endswith('.webp'):
                                        mime_type = 'image/webp'
                                
                                full_context.append(f"Filename: {file_name}")
                                full_context.append({
                                    'mime_type': mime_type,
                                    'data': content
                                })

                            # 3. Fallback: Try decoding as plain text
                            else:
                                try:
                                    text_content = content.decode('utf-8', errors='ignore')
                                    full_context.append(f"Filename: {file_name}\nContent:\n{text_content}")
                                except:
                                    full_context.append(f"Skipping file {file_name}: Unsupported format")
                        else:
                             full_context.append(f"Error reading file {file_name}: Download failed or content is empty.")
                            
                except Exception as sp_e:
                    logger.error(f"Error fetching SharePoint context: {sp_e}")
                    # We can't add error string to `full_context` now that it expects dicts/mixed. 
                    # If we really want to signal error to model, we could add a text part.
                    # full_context.append(f"Error reading file.") 

            response_text = llm_service.generate_response(message, history, full_context)

            return Response({"response": response_text}, status=status.HTTP_200_OK)
        except Exception as e:
            logger.error(f"Error in ChatView: {str(e)}")
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class SharePointFilesView(APIView):
    authentication_classes = [authentication.TokenAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        try:
            folder_id = request.query_params.get('folder_id')
            sp_service = SharePointService(user=request.user)
            files = sp_service.list_files(folder_id=folder_id)
            return Response({"files": files}, status=status.HTTP_200_OK)
        except Exception as e:
            logger.error(f"Error listing files: {str(e)}")
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
